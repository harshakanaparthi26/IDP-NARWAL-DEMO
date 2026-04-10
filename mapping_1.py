"""
interchange_mapper.py
─────────────────────
Maps interchange line items extracted from merchant statements
to your internal program catalog using:
  Phase 1 → TitanEmbeddings + FAISS  (fast candidate retrieval)
  Phase 2 → CrossEncoder reranker    (precision re-scoring)
  Phase 3 → Claude via Bedrock       (semantic mapping + confidence)

Built on your existing bedrock.py patterns.
"""

import json
import pickle
import numpy as np
import pandas as pd
import faiss
import boto3

from sentence_transformers import CrossEncoder

# ── Settings (mirror your settings.py pattern) ──────────────────────────────

class Settings:
    # Bedrock
    BEDROCK_MODEL_QA: str = "anthropic.claude-sonnet-4-5"          # Claude Sonnet
    EMBEDDING_MODEL: str  = "amazon.titan-embed-text-v2:0"
    TITAN_DIMENSIONS: int = 1024
    TITAN_NORMALIZE: bool = True

    # Reranker
    USE_RERANKER: bool    = True
    RERANKER_MODEL: str   = "cross-encoder/ms-marco-MiniLM-L-6-v2"

    # Pipeline tuning
    FAISS_TOP_K: int      = 20    # candidates from FAISS
    RERANKER_TOP_K: int   = 10    # candidates passed to Claude after reranking
    CONFIDENCE_THRESHOLD: float = 0.75   # below this → flagged for human review

    # Paths
    PROGRAM_SHEET_PATH: str  = "internal_programs.xlsx"   # your local Excel file
    FAISS_INDEX_PATH: str    = "program_index.faiss"
    METADATA_PATH: str       = "program_metadata.pkl"

settings = Settings()


# ── AWS / embedding clients (mirrors your bedrock.py pattern) ────────────────

session    = boto3.Session()
_s3        = session.client("s3")
_bedrock   = session.client("bedrock-runtime")
_emb       = None
_reranker  = None


# ── TitanEmbeddings (copied verbatim from your bedrock.py) ───────────────────

class TitanEmbeddings:
    """
    Bedrock Titan Embeddings V2 client.
    Model spec: amazon.titan-embed-text-v2:0
    """
    accept       = "application/json"
    content_type = "application/json"

    def __init__(
        self,
        model_id   = settings.EMBEDDING_MODEL,
        dimensions = settings.TITAN_DIMENSIONS,
        normalize  = settings.TITAN_NORMALIZE,
    ):
        self.client     = boto3.client("bedrock-runtime")
        self.model_id   = model_id
        self.dimensions = dimensions
        self.normalize  = normalize

    def embed(self, text: str):
        body = json.dumps({
            "inputText":  text,
            "dimensions": self.dimensions,
            "normalize":  self.normalize,
        })
        resp    = self.client.invoke_model(
            body=body,
            modelId=self.model_id,
            accept=self.accept,
            contentType=self.content_type,
        )
        payload = json.loads(resp["body"].read())
        return payload["embedding"]

    def embed_batch(self, texts):
        return [self.embed(t) for t in texts]

    # HF compatibility layer
    def embed_query(self, text: str):
        """Used during retrieval."""
        return self.embed(text)

    def embed_documents(self, texts):
        """Used during indexing & retrieval pool scoring."""
        return self.embed_batch(texts)


def get_embedding_model():
    global _emb
    if _emb is None:
        print(f"[RAG] Loading Titan embedding model: {settings.EMBEDDING_MODEL}")
        _emb = TitanEmbeddings(
            model_id   = settings.EMBEDDING_MODEL,
            dimensions = settings.TITAN_DIMENSIONS,
            normalize  = settings.TITAN_NORMALIZE,
        )
    return _emb


def get_reranker():
    global _reranker
    if _reranker is None and settings.USE_RERANKER:
        from sentence_transformers import CrossEncoder
        print(f"[RAG] Loading reranker: {settings.RERANKER_MODEL}")
        _reranker = CrossEncoder(settings.RERANKER_MODEL)
    return _reranker


# ── LLM call (mirrors your ask_on_text / _call pattern) ──────────────────────

def _call(
    system:      str,
    user:        str,
    model_id:    str   = None,
    max_tokens:  int   = 800,
    temperature: float = 0.0,
) -> str:
    """Raw Bedrock call — mirrors your existing _call() in bedrock.py."""
    model_id = model_id or settings.BEDROCK_MODEL_QA
    body = json.dumps({
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens":        max_tokens,
        "temperature":       temperature,
        "system":            system,
        "messages": [{"role": "user", "content": user}],
    })
    resp    = _bedrock.invoke_model(
        body=body,
        modelId=model_id,
        accept="application/json",
        contentType="application/json",
    )
    payload = json.loads(resp["body"].read())
    return payload["content"][0]["text"]


def ask_on_text(
    text:          str,
    prompt:        str,
    model_id:      str   = None,
    max_tokens:    int   = 800,
    temperature:   float = 0.0,
    answer_format: str   = "plain",   # "plain" | "json"
) -> str:
    """
    Send the entire text + a task prompt to the LLM in a single call.
    Mirrors your existing ask_on_text() in rag.py.
    """
    system = (
        "You are a careful assistant. Use ONLY the provided text. "
        "If something is missing, say 'insufficient context'. "
        "Be concise and accurate."
    )
    suffix = ""
    if answer_format == "json":
        suffix = "\n\nReturn valid JSON only. Do not include explanations."

    user = (
        f"=== TEXT START ===\n"
        f"{text}\n"
        f"=== TEXT END ===\n\n"
        f"=== TASK ===\n"
        f"{prompt}{suffix}"
    )
    return _call(system=system, user=user, model_id=model_id,
                 max_tokens=max_tokens, temperature=temperature)


# ── Program sheet helpers ─────────────────────────────────────────────────────

def load_program_sheet(path: str = settings.PROGRAM_SHEET_PATH) -> pd.DataFrame:
    """
    Load your internal program Excel/CSV.
    Creates a 'search_text' column that combines all relevant columns
    for embedding — adjust column names to match your actual sheet.
    """
    if path.endswith(".csv"):
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)

    # ── EDIT THESE to match your actual column names ──────────────────────────
    TEXT_COLUMNS = [
        "program_name",     # e.g. "CPS/RETAIL"
        "card_type",        # e.g. "Consumer Credit"
        "network",          # e.g. "Visa"
        "category",         # e.g. "Card Present"
        "sub_category",     # e.g. "Retail"
    ]
    # ─────────────────────────────────────────────────────────────────────────

    available = [c for c in TEXT_COLUMNS if c in df.columns]
    df["search_text"] = df[available].fillna("").astype(str).agg(" ".join, axis=1)
    df = df.reset_index(drop=True)
    print(f"[Sheet] Loaded {len(df)} programs from '{path}'")
    return df


# ── FAISS index build (run once, then save) ───────────────────────────────────

def build_and_save_index(
    df:         pd.DataFrame,
    index_path: str = settings.FAISS_INDEX_PATH,
    meta_path:  str = settings.METADATA_PATH,
):
    """
    Embed all rows with Titan and persist the FAISS index + metadata.
    Run this once whenever the program sheet changes.
    """
    emb    = get_embedding_model()
    texts  = df["search_text"].tolist()

    print(f"[Index] Embedding {len(texts)} programs via Titan (this takes ~2-3 min)...")
    vectors = np.array(emb.embed_batch(texts), dtype="float32")

    # Inner-product index (works as cosine sim because Titan normalises)
    index = faiss.IndexFlatIP(vectors.shape[1])
    index.add(vectors)

    faiss.write_index(index, index_path)
    with open(meta_path, "wb") as f:
        pickle.dump(df, f)

    print(f"[Index] Saved → {index_path}  |  metadata → {meta_path}")


def load_index(
    index_path: str = settings.FAISS_INDEX_PATH,
    meta_path:  str = settings.METADATA_PATH,
):
    """Load persisted FAISS index + program dataframe."""
    index = faiss.read_index(index_path)
    with open(meta_path, "rb") as f:
        df = pickle.load(f)
    print(f"[Index] Loaded {index.ntotal} vectors from '{index_path}'")
    return index, df


# ── Phase 1 — FAISS candidate retrieval ──────────────────────────────────────

def retrieve_candidates(
    query_text: str,
    index:      faiss.Index,
    df:         pd.DataFrame,
    top_k:      int = settings.FAISS_TOP_K,
) -> pd.DataFrame:
    emb       = get_embedding_model()
    vec       = np.array([emb.embed_query(query_text)], dtype="float32")
    distances, indices = index.search(vec, top_k)

    candidates = df.iloc[indices[0]].copy().reset_index(drop=True)
    candidates["faiss_score"] = distances[0]
    return candidates


# ── Phase 2 — CrossEncoder reranking ─────────────────────────────────────────

def rerank_candidates(
    query_text: str,
    candidates: pd.DataFrame,
    top_k:      int = settings.RERANKER_TOP_K,
) -> pd.DataFrame:
    reranker = get_reranker()
    if reranker is None:
        return candidates.head(top_k)

    pairs  = [(query_text, row["search_text"]) for _, row in candidates.iterrows()]
    scores = reranker.predict(pairs)

    candidates = candidates.copy()
    candidates["reranker_score"] = scores
    return (
        candidates
        .sort_values("reranker_score", ascending=False)
        .head(top_k)
        .reset_index(drop=True)
    )


# ── Phase 3 — Claude semantic mapping ────────────────────────────────────────

def map_with_claude(
    statement_item: dict,
    candidates:     pd.DataFrame,
) -> dict:
    """
    Uses ask_on_text() pattern — sends statement item + shortlisted candidates
    to Claude and gets back a structured JSON mapping result.
    """
    # Build a compact candidates table (only key columns)
    display_cols = [c for c in ["program_name", "card_type", "network", "rate", "category"]
                    if c in candidates.columns]
    candidates_table = candidates[display_cols].to_string()

    # The "text" is the full context; "prompt" is the task — mirrors ask_on_text()
    context = (
        f"STATEMENT ITEM:\n{json.dumps(statement_item, indent=2)}\n\n"
        f"INTERNAL PROGRAM CANDIDATES (index | details):\n{candidates_table}"
    )

    prompt = """You are matching interchange line items from a merchant statement
to an internal program catalog. Abbreviations may differ but represent the same concept.

Return a JSON object with EXACTLY these keys:
{
  "matched_index": <integer row index from candidates above>,
  "program_name":  "<matched program name>",
  "confidence":    <float 0.0–1.0>,
  "reasoning":     "<one sentence explanation>",
  "alt_index":     <second best index, or null>
}

Rules:
- confidence > 0.85 → very strong match
- confidence 0.75–0.85 → good match
- confidence < 0.75 → uncertain, needs human review
- If nothing fits, still return best guess with low confidence"""

    raw = ask_on_text(
        text          = context,
        prompt        = prompt,
        answer_format = "json",
        max_tokens    = 400,
        temperature   = 0.0,
    )

    # Strip any accidental markdown fences
    clean = raw.strip().replace("```json", "").replace("```", "").strip()
    return json.loads(clean)


# ── Full pipeline ─────────────────────────────────────────────────────────────

def process_statement(
    statement_items: list[dict],
    index:           faiss.Index,
    df:              pd.DataFrame,
    confidence_threshold: float = settings.CONFIDENCE_THRESHOLD,
) -> pd.DataFrame:
    """
    Maps a list of interchange line items from one merchant statement
    to the internal program catalog.

    Args:
        statement_items: List of dicts, each representing one interchange line.
                         e.g. [{"program_name": "CPS/RETAIL", "card_type": "Credit", ...}]
        index:           Loaded FAISS index.
        df:              Loaded program dataframe.
        confidence_threshold: Rows below this are flagged for human review.

    Returns:
        DataFrame with original columns + mapping results appended.
    """
    results = []

    for i, item in enumerate(statement_items):
        print(f"[Mapping] Item {i+1}/{len(statement_items)}: {item.get('program_name','?')}")

        # Build query text from item (same logic as row_to_text for the sheet)
        query_text = " ".join(str(v) for v in item.values() if v)

        # Phase 1 — FAISS
        candidates = retrieve_candidates(query_text, index, df)

        # Phase 2 — CrossEncoder rerank
        candidates = rerank_candidates(query_text, candidates)

        # Phase 3 — Claude
        try:
            mapping = map_with_claude(item, candidates)
        except (json.JSONDecodeError, KeyError) as e:
            print(f"  [WARN] Claude parse error: {e} — flagging for review")
            mapping = {
                "matched_index": 0,
                "program_name":  "PARSE_ERROR",
                "confidence":    0.0,
                "reasoning":     str(e),
                "alt_index":     None,
            }

        # Look up matched row to get the rate
        try:
            matched_row  = candidates.iloc[mapping["matched_index"]]
            internal_rate = matched_row.get("rate", "N/A")
            matched_name  = matched_row.get("program_name", mapping["program_name"])
        except (IndexError, KeyError):
            internal_rate = "N/A"
            matched_name  = mapping.get("program_name", "UNKNOWN")

        results.append({
            **item,
            "matched_program":  matched_name,
            "internal_rate":    internal_rate,
            "confidence":       mapping["confidence"],
            "reasoning":        mapping["reasoning"],
            "alt_match":        mapping.get("alt_index"),
            "needs_review":     mapping["confidence"] < confidence_threshold,
        })

    return pd.DataFrame(results)


def save_results_excel(results_df: pd.DataFrame, output_path: str = "mapping_results.xlsx"):
    """
    Save results to Excel with 'needs_review' rows highlighted in yellow.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results_df.to_excel(writer, index=False, sheet_name="Mappings")

        ws = writer.sheets["Mappings"]
        from openpyxl.styles import PatternFill
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Find the 'needs_review' column index (1-based for openpyxl)
        col_names = list(results_df.columns)
        # Header is row 1, data starts row 2
        for row_idx, needs_review in enumerate(results_df["needs_review"], start=2):
            if needs_review:
                for col_idx in range(1, len(col_names) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow

    print(f"[Output] Saved → {output_path}  "
          f"({results_df['needs_review'].sum()} rows flagged for review)")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os

    # ── STEP 1: Build index (only needed once, or when sheet changes) ─────────
    if not os.path.exists(settings.FAISS_INDEX_PATH):
        print("[Setup] Building FAISS index for the first time...")
        program_df = load_program_sheet(settings.PROGRAM_SHEET_PATH)
        build_and_save_index(program_df)
    else:
        print("[Setup] FAISS index already exists — skipping build.")

    # ── STEP 2: Load index ────────────────────────────────────────────────────
    index, program_df = load_index()

    # ── STEP 3: Load your extracted statement items ───────────────────────────
    # Replace this with however you load your extracted interchange data.
    # Each dict should have at minimum 'program_name', and whatever other
    # fields you extract (card_type, network, transaction_type, etc.)
    statement_items = [
        {"program_name": "CPS/RETAIL",     "card_type": "Consumer Credit", "network": "Visa"},
        {"program_name": "MC WLDT CR",     "card_type": "World Elite",     "network": "Mastercard"},
        {"program_name": "EIRF",           "card_type": "Consumer Credit", "network": "Visa"},
        {"program_name": "VS BASE II",     "card_type": "Consumer",        "network": "Visa"},
        {"program_name": "MC MERIT III",   "card_type": "Consumer Credit", "network": "Mastercard"},
        # ... load from your actual extraction output
    ]

    # ── STEP 4: Run pipeline ──────────────────────────────────────────────────
    results = process_statement(statement_items, index, program_df)

    # ── STEP 5: Save to Excel ─────────────────────────────────────────────────
    save_results_excel(results, "mapping_results.xlsx")

    # Quick summary
    total        = len(results)
    needs_review = results["needs_review"].sum()
    avg_conf     = results["confidence"].mean()
    print(f"\n── Summary ───────────────────────────────")
    print(f"  Total mapped   : {total}")
    print(f"  Auto-confirmed : {total - needs_review}  (confidence ≥ {settings.CONFIDENCE_THRESHOLD})")
    print(f"  Needs review   : {needs_review}")
    print(f"  Avg confidence : {avg_conf:.2f}")
    print(f"──────────────────────────────────────────")
